import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def compile_excel_buffer(pages_data: dict) -> bytes:
    """
    Compiles tabular JSON structures page-by-page into openpyxl sheets.
    Rules:
    - Dedicated worksheet sheet per page
    - Header formatting
    - Auto column widths
    """
    wb = Workbook()
    # Remove default worksheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    pages = pages_data.get("pages", [])
    if not pages:
        # Create a single default sheet if empty
        ws = wb.create_sheet(title="Empty Result")
        ws.cell(row=1, column=1, value="No tables extracted.")
    
    for page in pages:
        page_num = page.get("pageNumber", 1)
        headers = page.get("headers", [])
        rows = page.get("rows", [])

        sheet_title = f"Page {page_num}"
        ws = wb.create_sheet(title=sheet_title)

        # Style formats definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark slate
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        # 1. Write headers row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.row_dimensions[1].height = 26

        # 2. Write rows content
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=11)
                cell.border = thin_border
                ws.row_dimensions[row_idx].height = 20

        # 3. Dynamic Auto Column Width logic aligning characters length
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Add dynamic margin
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to a virtual binary buffer
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    excel_binary = virtual_file.getvalue()
    virtual_file.close()

    return excel_binary
